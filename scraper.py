"""
Trust & Safety / Risk job hunter.

Modes:
    python scraper.py daily    -> only NEW jobs since last run -> excel "Daily New" tab
    python scraper.py weekly   -> full current snapshot of every matching job -> excel "Weekly Snapshot" tab

Both modes regenerate docs/index.html, a live public page showing every
currently-open matching job, with newly-found roles marked. Serve it for
free via GitHub Pages pointed at the /docs folder.

State is kept in data/seen_jobs.json, committed back to the repo by the
GitHub Action after each run so "new since yesterday" works across runs.
"""

import os
import sys
import json
import time
import requests
import yaml
from datetime import datetime, timezone
from openpyxl import Workbook, load_workbook

CONFIG_PATH = "config.yaml"
STATE_PATH = "data/seen_jobs.json"
EXCEL_PATH = "data/jobs.xlsx"
HTML_PATH = "docs/index.html"
TIMEOUT = 15


def load_config():
    with open(CONFIG_PATH) as f:
        return yaml.safe_load(f)


def load_state():
    if os.path.exists(STATE_PATH):
        with open(STATE_PATH) as f:
            return json.load(f)
    return {}


def save_state(state):
    os.makedirs(os.path.dirname(STATE_PATH), exist_ok=True)
    with open(STATE_PATH, "w") as f:
        json.dump(state, f, indent=2)


def matches_keywords(title, keywords):
    title_lower = title.lower()
    return [kw for kw in keywords if kw.lower() in title_lower]


def fetch_greenhouse(token):
    url = f"https://boards-api.greenhouse.io/v1/boards/{token}/jobs"
    r = requests.get(url, timeout=TIMEOUT)
    r.raise_for_status()
    jobs = []
    for j in r.json().get("jobs", []):
        jobs.append({
            "id": str(j["id"]),
            "title": j["title"],
            "location": j.get("location", {}).get("name", ""),
            "url": j["absolute_url"],
        })
    return jobs


def fetch_lever(token):
    url = f"https://api.lever.co/v0/postings/{token}?mode=json"
    r = requests.get(url, timeout=TIMEOUT)
    r.raise_for_status()
    jobs = []
    for j in r.json():
        jobs.append({
            "id": j["id"],
            "title": j["text"],
            "location": j.get("categories", {}).get("location", ""),
            "url": j["hostedUrl"],
        })
    return jobs


def fetch_ashby(token):
    url = f"https://api.ashbyhq.com/posting-api/job-board/{token}"
    r = requests.get(url, timeout=TIMEOUT)
    r.raise_for_status()
    jobs = []
    for j in r.json().get("jobs", []):
        jobs.append({
            "id": j["id"],
            "title": j["title"],
            "location": j.get("location", ""),
            "url": j["jobUrl"],
        })
    return jobs


def fetch_custom(company):
    # Custom career sites vary too much for one generic scraper to handle
    # reliably (JS-rendered pages, different HTML structures, etc).
    # This is a placeholder that companies of type "custom" hit --
    # see README.md "Custom sites" section for how to fill these in
    # (e.g. with Playwright) once you know each site's structure.
    return []


def fetch_company_jobs(company):
    ctype = company["type"]
    try:
        if ctype == "greenhouse":
            return fetch_greenhouse(company["token"])
        elif ctype == "lever":
            return fetch_lever(company["token"])
        elif ctype == "ashby":
            return fetch_ashby(company["token"])
        elif ctype == "custom":
            return fetch_custom(company)
    except Exception as e:
        print(f"  [warn] {company['name']}: fetch failed ({e})")
        return []
    return []


def collect_matches(config):
    """Returns list of dicts: company, title, location, url, matched_keywords, job_key"""
    all_matches = []
    for company in config["companies"]:
        jobs = fetch_company_jobs(company)
        for job in jobs:
            hits = matches_keywords(job["title"], config["keywords"])
            if hits:
                all_matches.append({
                    "company": company["name"],
                    "title": job["title"],
                    "location": job.get("location", ""),
                    "url": job["url"],
                    "matched_keywords": ", ".join(hits),
                    "job_key": f"{company['name']}::{job['id']}",
                })
        time.sleep(0.3)  # be polite to the APIs
    return all_matches


def write_excel(matches, path, mode):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    if os.path.exists(path):
        wb = load_workbook(path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    sheet_name = "Weekly Snapshot" if mode == "weekly" else "Daily New"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name, 0)
    ws.append(["Date Found", "Company", "Title", "Location", "Matched Keywords", "URL"])

    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    for m in matches:
        ws.append([today, m["company"], m["title"], m["location"], m["matched_keywords"], m["url"]])

    for col in ws.columns:
        max_len = max(len(str(c.value)) for c in col if c.value)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    wb.save(path)


def write_html(matches, path):
    os.makedirs(os.path.dirname(path), exist_ok=True)

    companies = sorted({m["company"] for m in matches})
    keywords_flat = sorted({kw.strip() for m in matches for kw in m["matched_keywords"].split(",")})
    new_count = sum(1 for m in matches if m.get("is_new"))
    scanned_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    data_json = json.dumps([
        {
            "company": m["company"],
            "title": m["title"],
            "location": m["location"],
            "keywords": [k.strip() for k in m["matched_keywords"].split(",")],
            "url": m["url"],
            "found": m["date_found"],
            "isNew": bool(m.get("is_new")),
        }
        for m in matches
    ])

    company_options = "".join(f'<option value="{c}">{c}</option>' for c in companies)
    keyword_options = "".join(f'<option value="{k}">{k}</option>' for k in keywords_flat)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Signal &mdash; T&amp;S / Risk role tracker</title>
<style>
  :root {{
    --ink: #0F1620;
    --panel: #16202C;
    --line: #263447;
    --text: #D9DFE6;
    --text-dim: #8695A7;
    --signal: #E8A94C;
    --tag: #5FA8A0;
    --focus: #7FB8E0;
  }}
  * {{ box-sizing: border-box; }}
  body {{
    margin: 0;
    background: var(--ink);
    color: var(--text);
    font-family: 'IBM Plex Sans', system-ui, sans-serif;
    -webkit-font-smoothing: antialiased;
  }}
  code, .mono, td.location, td.found, .count-num {{
    font-family: 'IBM Plex Mono', ui-monospace, Menlo, monospace;
  }}
  header {{
    padding: 2.5rem 2rem 1.5rem;
    border-bottom: 1px solid var(--line);
    max-width: 1200px;
    margin: 0 auto;
  }}
  .eyebrow {{
    display: flex;
    align-items: center;
    gap: 0.5rem;
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.75rem;
    letter-spacing: 0.08em;
    text-transform: uppercase;
    color: var(--text-dim);
    margin-bottom: 0.75rem;
  }}
  .dot {{
    width: 8px; height: 8px; border-radius: 50%;
    background: var(--signal);
    animation: pulse 2.4s ease-in-out infinite;
  }}
  @media (prefers-reduced-motion: reduce) {{ .dot {{ animation: none; }} }}
  @keyframes pulse {{
    0%, 100% {{ opacity: 1; box-shadow: 0 0 0 0 rgba(232,169,76,0.5); }}
    50% {{ opacity: 0.6; box-shadow: 0 0 0 5px rgba(232,169,76,0); }}
  }}
  h1 {{
    font-size: 1.9rem;
    margin: 0 0 0.4rem;
    font-weight: 600;
    letter-spacing: -0.01em;
  }}
  .sub {{ color: var(--text-dim); font-size: 0.95rem; max-width: 60ch; }}
  .counts {{
    display: flex;
    gap: 2rem;
    margin-top: 1.5rem;
    flex-wrap: wrap;
  }}
  .count-block .count-num {{
    font-size: 1.6rem;
    color: var(--focus);
    display: block;
  }}
  .count-block .count-label {{
    font-size: 0.75rem;
    color: var(--text-dim);
    text-transform: uppercase;
    letter-spacing: 0.06em;
  }}
  main {{ max-width: 1200px; margin: 0 auto; padding: 1.5rem 2rem 3rem; }}
  .controls {{
    display: flex;
    gap: 0.75rem;
    flex-wrap: wrap;
    margin-bottom: 1.25rem;
  }}
  .controls input, .controls select {{
    background: var(--panel);
    border: 1px solid var(--line);
    color: var(--text);
    padding: 0.55rem 0.75rem;
    border-radius: 6px;
    font-size: 0.9rem;
    font-family: inherit;
  }}
  .controls input {{ flex: 1; min-width: 200px; }}
  .controls input:focus, .controls select:focus, th button:focus-visible {{
    outline: 2px solid var(--focus);
    outline-offset: 1px;
  }}
  table {{
    width: 100%;
    border-collapse: collapse;
    font-size: 0.9rem;
  }}
  thead th {{
    text-align: left;
    padding: 0.6rem 0.8rem;
    border-bottom: 1px solid var(--line);
    color: var(--text-dim);
    font-weight: 500;
    font-size: 0.78rem;
    text-transform: uppercase;
    letter-spacing: 0.04em;
  }}
  thead th button {{
    background: none; border: none; color: inherit; font: inherit;
    cursor: pointer; padding: 0; text-transform: inherit; letter-spacing: inherit;
  }}
  tbody tr {{ border-bottom: 1px solid var(--line); }}
  tbody tr:hover {{ background: var(--panel); }}
  td {{ padding: 0.7rem 0.8rem; vertical-align: top; }}
  td.title a {{ color: var(--text); text-decoration: none; font-weight: 500; }}
  td.title a:hover {{ color: var(--focus); text-decoration: underline; }}
  td.location, td.found {{ color: var(--text-dim); white-space: nowrap; }}
  .badge-new {{
    display: inline-block;
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.65rem;
    color: var(--ink);
    background: var(--signal);
    padding: 0.1rem 0.4rem;
    border-radius: 4px;
    margin-left: 0.5rem;
    vertical-align: middle;
  }}
  .kw {{
    display: inline-block;
    font-size: 0.72rem;
    color: var(--tag);
    border: 1px solid rgba(95,168,160,0.4);
    padding: 0.05rem 0.4rem;
    border-radius: 999px;
    margin: 0.1rem 0.2rem 0.1rem 0;
  }}
  .empty {{ color: var(--text-dim); padding: 3rem 0; text-align: center; }}
  footer {{
    max-width: 1200px; margin: 0 auto; padding: 1.5rem 2rem 3rem;
    color: var(--text-dim); font-size: 0.8rem;
  }}
  footer a {{ color: var(--tag); }}
</style>
</head>
<body>
<header>
  <div class="eyebrow"><span class="dot" aria-hidden="true"></span> live &middot; auto-updates on schedule</div>
  <h1>Signal</h1>
  <p class="sub">Tracking Trust &amp; Safety, Risk, and Integrity roles across {len(companies)} companies. Last scanned {scanned_at}.</p>
  <div class="counts">
    <div class="count-block"><span class="count-num">{len(matches)}</span><span class="count-label">open matches</span></div>
    <div class="count-block"><span class="count-num">{new_count}</span><span class="count-label">new this run</span></div>
    <div class="count-block"><span class="count-num">{len(companies)}</span><span class="count-label">companies w/ matches</span></div>
  </div>
</header>
<main>
  <div class="controls">
    <input id="search" type="text" placeholder="Search title or company&hellip;" aria-label="Search jobs">
    <select id="companyFilter" aria-label="Filter by company">
      <option value="">All companies</option>
      {company_options}
    </select>
    <select id="keywordFilter" aria-label="Filter by matched keyword">
      <option value="">All keywords</option>
      {keyword_options}
    </select>
  </div>
  <table id="jobTable">
    <thead>
      <tr>
        <th><button data-sort="company">Company</button></th>
        <th><button data-sort="title">Title</button></th>
        <th><button data-sort="location">Location</button></th>
        <th>Matched</th>
        <th><button data-sort="found">Found</button></th>
      </tr>
    </thead>
    <tbody id="jobBody"></tbody>
  </table>
  <div id="emptyState" class="empty" hidden>No roles match your filters right now.</div>
</main>
<footer>
  Generated automatically &middot; source: <a href="https://github.com" target="_blank" rel="noopener">GitHub Action</a> &middot; data refreshes on the schedule set in the repo's workflows.
</footer>
<script>
  const DATA = {data_json};
  let sortKey = 'found';
  let sortDir = -1;

  const $ = (sel) => document.querySelector(sel);
  const body = $('#jobBody');
  const empty = $('#emptyState');

  function render() {{
    const q = $('#search').value.trim().toLowerCase();
    const co = $('#companyFilter').value;
    const kw = $('#keywordFilter').value;

    let rows = DATA.filter(j => {{
      const matchesQ = !q || j.title.toLowerCase().includes(q) || j.company.toLowerCase().includes(q);
      const matchesCo = !co || j.company === co;
      const matchesKw = !kw || j.keywords.includes(kw);
      return matchesQ && matchesCo && matchesKw;
    }});

    rows.sort((a, b) => {{
      const av = (a[sortKey] || '').toString().toLowerCase();
      const bv = (b[sortKey] || '').toString().toLowerCase();
      if (av < bv) return -1 * sortDir;
      if (av > bv) return 1 * sortDir;
      return 0;
    }});

    body.innerHTML = '';
    empty.hidden = rows.length !== 0;

    for (const j of rows) {{
      const tr = document.createElement('tr');
      tr.innerHTML = `
        <td>${{j.company}}</td>
        <td class="title"><a href="${{j.url}}" target="_blank" rel="noopener">${{j.title}}</a>${{j.isNew ? '<span class="badge-new">NEW</span>' : ''}}</td>
        <td class="location">${{j.location || '&mdash;'}}</td>
        <td>${{j.keywords.map(k => `<span class="kw">${{k}}</span>`).join('')}}</td>
        <td class="found">${{j.found}}</td>
      `;
      body.appendChild(tr);
    }}
  }}

  $('#search').addEventListener('input', render);
  $('#companyFilter').addEventListener('change', render);
  $('#keywordFilter').addEventListener('change', render);
  document.querySelectorAll('th button').forEach(btn => {{
    btn.addEventListener('click', () => {{
      const key = btn.dataset.sort;
      if (sortKey === key) {{ sortDir *= -1; }} else {{ sortKey = key; sortDir = 1; }}
      render();
    }});
  }});

  render();
</script>
</body>
</html>
"""
    with open(path, "w") as f:
        f.write(html)


def run(mode):
    config = load_config()
    state = load_state()
    all_matches = collect_matches(config)

    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    # figure out which matches are new BEFORE mutating state, so the html
    # page can badge them regardless of which mode we're running
    new_keys = {m["job_key"] for m in all_matches if m["job_key"] not in state}
    for m in all_matches:
        m["is_new"] = m["job_key"] in new_keys
        m["date_found"] = state.get(m["job_key"], today)

    if mode == "daily":
        new_matches = [m for m in all_matches if m["is_new"]]
        for m in new_matches:
            state[m["job_key"]] = today
        save_state(state)
        write_excel(new_matches, EXCEL_PATH, "daily")
        write_html(all_matches, HTML_PATH)
        print(f"Found {len(new_matches)} NEW matching jobs. Written to {EXCEL_PATH} (tab: 'Daily New') and {HTML_PATH}.")
    elif mode == "weekly":
        # weekly = full snapshot, regardless of "seen" status
        for m in all_matches:
            state.setdefault(m["job_key"], today)
        save_state(state)
        write_excel(all_matches, EXCEL_PATH, "weekly")
        write_html(all_matches, HTML_PATH)
        print(f"Found {len(all_matches)} matching jobs (full snapshot). Written to {EXCEL_PATH} (tab: 'Weekly Snapshot') and {HTML_PATH}.")
    else:
        print("Usage: python scraper.py [daily|weekly]")
        sys.exit(1)


if __name__ == "__main__":
    mode = sys.argv[1] if len(sys.argv) > 1 else "daily"
    run(mode)
